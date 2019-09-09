from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from bs4 import BeautifulSoup as soup
import xlsxwriter as xl
from openpyxl import load_workbook
import openpyxl
import time
import re
from os import system, name
League_links ={
"JP1":"https://www.flashscore.com/football/japan/j-league/results/",
"CHI1":"https://www.flashscore.com/football/china/super-league/results/",
"SK1":"https://www.flashscore.com/football/south-korea/k-league-1/results/",
"SWD1":"https://www.flashscore.com/football/sweden/allsvenskan/results/",
"FIN1":"https://www.flashscore.com/football/finland/veikkausliiga/results/",
"BRZ1":"https://www.flashscore.com/football/brazil/serie-a/results/",
"RUS1":"https://www.flashscore.com/football/russia/premier-league/results/",
"RUS2":"https://www.flashscore.com/football/russia/fnl/results/",
"BEL1":"https://www.flashscore.com/football/belgium/jupiler-league/results/",
"BUL1":"https://www.flashscore.com/football/bulgaria/parva-liga/results/",
"CZR1":"https://www.flashscore.com/football/czech-republic/1-liga/results/",
"NOR1":"https://www.flashscore.com/football/norway/eliteserien/results/",
"RMN1":"https://www.flashscore.com/football/romania/liga-1/results/",
"SRB1":"https://www.flashscore.com/football/serbia/super-liga/results/",
"ARG1":"https://www.flashscore.com/football/argentina/superliga/results/",
"CHL1":"https://www.flashscore.com/football/chile/primera-division/results/",
"COL1":"https://www.flashscore.com/football/colombia/liga-aguila/results/",
}
def main():
	league_name = input("Enter League Name: ")
	getAll = input("Do you need Every Week?: (yes/no) ")
	if getAll=="yes": getAll = True
	else: getAll = False
	file_name = league_name + '/' + league_name + '_ID.xlsx'
	my_soup = get_soup(League_links[league_name],getAll)
	round_nums, ids = get_id(my_soup)
	write_to_xlsx(file_name, round_nums, ids)
	print(round_nums)
#------------------------------------------------------------------#
def get_soup(url,getAll):
	options = Options()
	options.add_argument("--headless")
	driver = webdriver.Firefox(firefox_options=options,
			   executable_path = '/usr/local/bin/geckodriver')
	driver.get(url)
	if(getAll):
		try: button = driver.find_element_by_xpath('//*[@id="live-table"]/div/div/div/a')
		except: pass
		for i in range(10):
			time.sleep(5)
			try: button.click()
			except: print("Failed")
	html = driver.page_source
	driver.quit()
	my_soup = soup(html,"html.parser")
	return my_soup
#------------------------------------------------------------------#
def get_id(my_soup):
	table = my_soup.find('div',{'class':'soccer'})
	Round = ""
	round_nums = []
	ids = []
	for div in table (re.compile('^div')):
		strDiv = str(div)
		if("id=" in strDiv and Round != ""):
			begin = strDiv.rfind('id="') + 8
			end = strDiv.rfind('title="') - 2
			matchID = strDiv[begin:end]
			ids.append((Round, matchID))
		if("Round" in strDiv):
			begin = strDiv.rfind('c">') + 3
			end = strDiv.rfind('<')
			Round = strDiv[begin:end]
			round_nums.append(Round)
	return round_nums, ids
#------------------------------------------------------------------#
def write_to_xlsx(file_name, round_nums, ids):
	id_book = xl.Workbook(file_name)
	new_round = []
	for i in range(len(round_nums)):
		try:
			id_book.add_worksheet(round_nums[i])
			new_round.append(round_nums[i])
		except:
		    	pass
	id_book.close()
	for i in range(len(new_round)):
		wb = load_workbook(file_name)
		ws = wb[(new_round[i])]
		row = 1
		col = 1
		for r in range(len(ids)):
		    
			if(ids[r][0] == new_round[i]):
				ws.cell(row,col).value = ('https://www.soccer24.com/match/'+ids[r][1]+'/')
				row += 1
		else:
			pass
		wb.save(file_name)

main()
